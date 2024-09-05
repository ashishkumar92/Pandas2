from openpyxl import Workbook # pip install openpyxl
from openpyxl.styles import Font, Color, colors
from openpyxl.utils import get_column_letter
 # see example 1 or
listA=['DistanceA (m)',1,2,3,4,5,6,7,8,8,9,9,1,2,3,4,5,6,7,8,8,9,9,1,2,3,4,5,6,7,8,8,9,9]
listB=['DistanceB (m)',1,2,3,4,5,6,7,8,8,9,9,1,2,3,4,5,6,7,8,8,9,9,1,2,3,4,5,6,7,8,8,9,9]
L=[listA, listB,listA, listB,listA, listB,listA, listB,listA, listB,listA, listB,listA, listB,listA, listB]
wb = Workbook()
ws1 = wb.active # work sheet
ws1.title = "Pyxl"
ft_1 = Font(name='Calibri',color=colors.BLUE, bold=True, size=11,underline='none')
ft_2 = Font(name='Calibri',color=colors.BLUE, bold=True, size=11,underline='none')
ft_3 = Font(name='Calibri',color=colors.BLACK, bold=True, size=11,underline='none')
ft_4 = Font(name='Calibri',color=colors.BLACK, bold=True, size=11,underline='none')

for i in range(len(listA)):
    for j in range (len(L)):
        _ = ws1.cell(column=j+1, row=i+1, value=L[j][i])
        if i==len(listA)-1 :
            _ = ws1.cell(column=j+1, row=i+3, value='=SUM('+get_column_letter(j+2)+str(i-len(listA)+2)+':'+ get_column_letter(j+2)+str(i+2)+')')
            _.font=ft_1
            _ = ws1.cell(column=j+1, row=i+4, value='=average('+get_column_letter(j+2)+str(i-len(listA)+2)+':'+ get_column_letter(j+2)+str(i+2)+')')
            _.font=ft_2
            _.number_format='#,#0.0'
            _ = ws1.cell(column=j+1, row=i+5, value='=max('+get_column_letter(j+2)+str(i-len(listA)+2)+':'+ get_column_letter(j+2)+str(i+2)+')')
            _.font=ft_3
    # ws1.append([listA[row]])

for i, name in enumerate(['Summation', 'Average', 'Maximum']):
    _ = ws1.cell(column=1, row=len(listA)+2+i, value=name)
    _.font=ft_4




wb.save('xlfile1.xlsx')